from io import BytesIO

from django.db.models import F, Func
from openpyxl import Workbook
from openpyxl.styles import Font


class SearchExcelExporter:
    BASE_COLUMNS = ["resourceinstanceid", "graph_slug"]

    def export(self, queryset, *, language=None):
        languages = [language] if language else self._collect_languages(queryset)
        descriptor_headers = [
            col for lang in languages for col in (f"{lang}-name", f"{lang}-description")
        ]
        columns = self.BASE_COLUMNS + descriptor_headers

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Search Results"

        header_font = Font(bold=True)
        for column_index, header in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.font = header_font

        for row_index, resource in enumerate(
            queryset.select_related("graph").iterator(), start=2
        ):
            worksheet.cell(
                row=row_index,
                column=1,
                value=str(resource.resourceinstanceid),
            )
            worksheet.cell(
                row=row_index,
                column=2,
                value=resource.graph.slug if resource.graph else "",
            )
            for lang_index, lang in enumerate(languages):
                descriptor = (resource.descriptors or {}).get(lang)
                name = descriptor.get("name", "") if descriptor else ""
                description = descriptor.get("description", "") if descriptor else ""
                if name == "Undefined":
                    name = ""
                if description == "Undefined":
                    description = ""

                column_index = 3 + lang_index * 2
                worksheet.cell(row=row_index, column=column_index, value=name)
                worksheet.cell(
                    row=row_index, column=column_index + 1, value=description
                )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _collect_languages(queryset):
        return sorted(
            queryset.filter(descriptors__isnull=False)
            .annotate(lang=Func(F("descriptors"), function="JSONB_OBJECT_KEYS"))
            .values_list("lang", flat=True)
            .distinct()
        )
