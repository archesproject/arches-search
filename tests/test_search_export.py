from unittest.mock import MagicMock, patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from arches_search.utils.search_export import SearchExcelExporter

# python manage.py test tests.test_search_export --settings="tests.test_settings"


class SearchExcelExporterTest(SimpleTestCase):
    RESOURCE1_ID = "d5c7e72e-4805-4f32-90d7-0db93176b2c4"
    RESOURCE2_ID = "db91f6fd-e70d-44e1-9492-00b87018cd78"

    def _make_resource(self, resourceinstanceid, graph_slug, descriptors=None):
        resource = MagicMock()
        resource.resourceinstanceid = resourceinstanceid
        resource.graph.slug = graph_slug
        resource.descriptors = descriptors
        return resource

    def _export_with_resources(self, resources, languages, language=None):
        queryset = MagicMock()
        queryset.select_related.return_value.iterator.return_value = resources
        exporter = SearchExcelExporter()
        with patch.object(
            SearchExcelExporter, "_collect_languages", return_value=languages
        ):
            return exporter.export(queryset, language=language)

    def _get_row(self, output, row_index):
        ws = load_workbook(filename=output).active
        return [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]

    def test_get_descriptors_of_system_lang(self):
        resource = self._make_resource(
            self.RESOURCE1_ID,
            "heritage_site",
            descriptors={
                "en": {"name": "Parthenon", "description": "An ancient temple"},
                "fr": {"name": "Parthénon", "description": "Un temple antique"},
            },
        )
        # Passing language="en" should skip _collect_languages and only produce en columns
        output = self._export_with_resources([resource], languages=[], language="en")
        headers = self._get_row(output, row_index=1)
        row = self._get_row(output, row_index=2)

        self.assertEqual(
            headers, ["resourceinstanceid", "graph_slug", "en-name", "en-description"]
        )
        self.assertEqual(row[2], "Parthenon")
        self.assertEqual(row[3], "An ancient temple")

    def test_get_all_descriptors(self):
        resource_en_only = self._make_resource(
            self.RESOURCE1_ID,
            "heritage_site",
            descriptors={"en": {"name": "Parthenon", "description": "A temple"}},
        )
        resource_fr_only = self._make_resource(
            self.RESOURCE2_ID,
            "heritage_site",
            descriptors={"fr": {"name": "Parthénon", "description": "Un temple"}},
        )
        output = self._export_with_resources(
            [resource_en_only, resource_fr_only], ["en", "fr"]
        )
        headers = self._get_row(output, row_index=1)
        row1 = self._get_row(output, row_index=2)
        row2 = self._get_row(output, row_index=3)

        self.assertEqual(
            headers[2:], ["en-name", "en-description", "fr-name", "fr-description"]
        )
        self.assertEqual(row1[2], "Parthenon")
        self.assertEqual(row1[3], "A temple")
        self.assertFalse(row1[4])  # fr-name missing for resource 1
        self.assertFalse(row1[5])  # fr-description missing for resource 1
        self.assertFalse(row2[2])  # en-name missing for resource 2
        self.assertFalse(row2[3])  # en-description missing for resource 2
        self.assertEqual(row2[4], "Parthénon")
        self.assertEqual(row2[5], "Un temple")
